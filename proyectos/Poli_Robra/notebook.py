import requests
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# URL del servidor ESP32
url = "http://192.168.123.43/"  # Dirección IP del AP de la ESP32

def get_temperature():
    try:
        # Enviar una solicitud GET al servidor
        response = requests.get(url)
        
        # Verificar si la solicitud fue exitosa
        if response.status_code == 200:
            # Imprimir el texto de la respuesta (datos de temperatura)
            print("Datos de temperatura recibidos del ESP32:")
            print(response.text)
            
            # Extraer los datos de temperatura del texto de la respuesta
            temp1 = None
            temp2 = None
            for line in response.text.split('\n'):
                if "Temp1:" in line:
                    temp1 = float(line.split(":")[1].strip().split()[0])
                if "Temp2:" in line:
                    temp2 = float(line.split(":")[1].strip().split()[0])
            
            # Verificar que ambos datos de temperatura se hayan extraído correctamente
            if temp1 is not None and temp2 is not None:
                return temp1, temp2
            else:
                print("Error: Datos de temperatura incompletos.")
                return None, None
        else:
            print(f"Error al recuperar los datos. Código de estado HTTP: {response.status_code}")
            return None, None
    except requests.exceptions.RequestException as e:
        print(f"Ocurrió un error: {e}")
        return None, None

def write_to_excel(hora, temp1, temp2):
    # Generar el nombre del archivo basado en la fecha actual
    current_date = datetime.now().strftime("%d_%m_%y")
    file_path = f"D:/DHT11/ESP_{current_date}.xlsx"
    
    # Verificar si la carpeta existe
    if not os.path.exists("D:/DHT11/"):
        print("La carpeta D:/DHT11/ no existe. Creándola...")
        os.makedirs("D:/DHT11/")
    
    try:
        # Intentar cargar el archivo existente
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo
        workbook = Workbook()
        sheet = workbook.active
        # Escribir los encabezados
        sheet.append(["Hora", "Temp1", "Temp2"])
    
    # Agregar los datos con la hora actual
    sheet.append([hora, temp1, temp2])
    
    # Guardar el archivo
    workbook.save(file_path)
    print(f"Datos guardados en {file_path}")

# Bucle principal para obtener datos cada 3 minutos
while True:
    temp1, temp2 = get_temperature()
    if temp1 is not None and temp2 is not None:
        hora_actual = datetime.now().strftime("%H:%M:%S")  # Solo la hora
        write_to_excel(hora_actual, temp1, temp2)
    # Esperar 3 minutos antes de la próxima solicitud
    time.sleep(180)
